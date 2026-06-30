# -*- coding: utf-8 -*-
"""
선사 등 다수 수신자에게 뉴스레터를 'BCC 한 통'으로 보내는 도구.

흐름:
  1) ../../recipients/ 폴더의 CSV(또는 XLSX/TXT)에서 이메일을 '패턴으로' 전부 추출
     - 열 위치/제목 상관없이 a@b.c 형태면 모두 인식, 대소문자 무시 중복 제거
  2) 같은 폴더의 create_outlook_email.py 의 HTML·이미지·제목을 그대로 재사용
  3) 수신자 전원을 BCC 에 넣은 메일 1통을 Outlook 새 창으로 띄움 (자동 발송 X)
     - 창에서 인원수/본문 확인 후 직접 '보내기'

엑셀(.xlsx)을 그대로 쓰려면 openpyxl 이 필요합니다(run.bat 자동 설치).
가장 간단한 방법: 엑셀에서 '다른 이름으로 저장 → CSV' 후 그 CSV 를 recipients/ 에 넣기.
"""
import os
import re
import glob

import create_outlook_email as base  # 같은 폴더의 단일 발송 스크립트 재사용

BASE = os.path.dirname(os.path.abspath(__file__))
# 저장소 공용 수신자 폴더 (issues/<호>/ 기준 두 단계 위)
REC_DIR = os.path.normpath(os.path.join(BASE, "..", "..", "recipients"))

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# 한 통 수신자 권장 상한(정책별 상이, Exchange Online 기본 500). 넘으면 경고.
MAX_PER_MESSAGE = 450


def _read_text(path):
    """CSV/TXT 를 인코딩 자동 판별로 통째로 읽어 문자열 반환."""
    for enc in ("utf-8-sig", "cp949", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    return ""


def _read_xlsx(path):
    """XLSX 의 모든 셀 값을 줄바꿈으로 이어붙여 반환 (openpyxl 필요)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [건너뜀] openpyxl 미설치로 xlsx 를 못 읽음:", os.path.basename(path))
        print("           해결: py -m pip install openpyxl   (또는 엑셀에서 CSV 로 저장)")
        return ""
    wb = load_workbook(path, read_only=True, data_only=True)
    buf = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    buf.append(str(cell))
    return "\n".join(buf)


def load_recipients():
    """recipients/ 의 모든 CSV/XLSX/TXT 에서 이메일을 추출·중복제거해 리스트 반환."""
    if not os.path.isdir(REC_DIR):
        print("[오류] 수신자 폴더가 없습니다:", REC_DIR)
        print("       이 폴더를 만들고 그 안에 CSV(또는 XLSX) 파일을 넣어주세요.")
        return []

    files = [p for p in glob.glob(os.path.join(REC_DIR, "*"))
             if p.lower().endswith((".csv", ".txt", ".xlsx"))
             # ~$… : 엑셀 임시 잠금파일 / _… : 샘플·제외용(읽지 않음)
             and not os.path.basename(p).startswith(("~$", "_"))]
    if not files:
        print("[오류] recipients 폴더에 CSV/XLSX/TXT 파일이 없습니다:", REC_DIR)
        return []

    emails, seen, dup = [], set(), 0
    for path in sorted(files):
        name = os.path.basename(path)
        text = _read_xlsx(path) if path.lower().endswith(".xlsx") else _read_text(path)
        found = EMAIL_RE.findall(text)
        print(f"  읽음: {name}  (이메일 {len(found)}개 발견)")
        for e in found:
            key = e.lower()
            if key in seen:
                dup += 1
                continue
            seen.add(key)
            emails.append(e)

    print(f"\n[집계] 고유 이메일 {len(emails)}명  (중복 {dup}개 제거)")
    return emails


def main():
    print("=== 대량 BCC 발송 (검토 후 직접 '보내기') ===")
    print("수신자 폴더:", REC_DIR)
    print()

    emails = load_recipients()
    if not emails:
        return

    head = ", ".join(emails[:5])
    print("  예시:", head, ("..." if len(emails) > 5 else ""))
    if len(emails) > MAX_PER_MESSAGE:
        print(f"\n[경고] 수신자 {len(emails)}명 — 한 통 권장 상한({MAX_PER_MESSAGE})을 넘습니다.")
        print("       Exchange 정책에 따라 거부될 수 있습니다. 목록을 나눠 보내는 것을 권장합니다.")

    if not os.path.exists(base.HTML_PATH):
        print("[오류] 본문 HTML 을 찾을 수 없습니다:", base.HTML_PATH)
        return

    import win32com.client
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # 0 = olMailItem
    mail.Subject = base.SUBJECT
    mail.BCC = "; ".join(emails)   # 250명을 BCC 로 → 서로 주소가 안 보임
    # 일부 Exchange 는 To 가 비어 있으면 거부합니다. 그럴 땐 아래 주석을 풀고
    # 본인(발신) 주소를 넣으세요:
    # mail.To = "me@avikus.ai"

    mail.HTMLBody = base.load_html()
    n = base.attach_inline_images(mail)
    print(f"\n  인라인 이미지 {n}개 삽입")

    mail.Display(False)  # 창만 열기 (발송 X). 검토 후 직접 '보내기'
    print("\n완료: Outlook 새 메일 창이 열렸습니다.")
    print("  ① BCC 인원수 확인  ② 열려 있던 옛 초안창 모두 닫기")
    print("  ③ 이 창에서만 '보내기'  ④ 보낸 뒤 '보낼 편지함(Outbox)'이 비었는지 확인")


if __name__ == "__main__":
    main()
